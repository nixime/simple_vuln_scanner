import copy
import re
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.formula.translate import Translator
from helpers.cvss_helper import CVSSHelper

class ExcelHelper:
    
    @staticmethod
    def copy_data_validations(source_ws, dest_ws):
        """Clones data validations from source worksheet to destination."""
        if not source_ws.data_validations:
            return
        for dv in source_ws.data_validations.dataValidation:
            new_dv = copy.copy(dv)
            dest_ws.add_data_validation(new_dv)

    @staticmethod
    def apply_formatting_to_range(worksheet, source_row_idx, start_row, count):
        """Clones styles from a source row to a target range."""
        if start_row <= source_row_idx or count <= 0:
            return

        max_col = worksheet.max_column
        row_styles = []
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=source_row_idx, column=col)
            row_styles.append({
                'font': copy.copy(cell.font),
                'border': copy.copy(cell.border),
                'fill': copy.copy(cell.fill),
                'number_format': cell.number_format,
                'protection': copy.copy(cell.protection),
                'alignment': copy.copy(cell.alignment)
            })

        for r_idx in range(start_row, start_row + count):
            for c_idx in range(1, max_col + 1):
                target_cell = worksheet.cell(row=r_idx, column=c_idx)
                s = row_styles[c_idx - 1]
                target_cell.font, target_cell.border = s['font'], s['border']
                target_cell.fill, target_cell.number_format = s['fill'], s['number_format']
                target_cell.protection, target_cell.alignment = s['protection'], s['alignment']

    @staticmethod
    def apply_data_validation_rules(worksheet, source_row_idx, count):
        """Extends data validation ranges to cover the newly added rows."""
        if not worksheet.data_validations:
            return
        max_col = worksheet.max_column
        for c in range(1, max_col + 1):
            column_letter = get_column_letter(c)
            target_range = f"{column_letter}{source_row_idx}:{column_letter}{source_row_idx+count}"

            for dv in worksheet.data_validations.dataValidation:
                for r in dv.ranges:
                    range_str = r.coord
                    if ":" not in range_str: range_str = f"{range_str}:{range_str}"
                    min_col, min_row, max_col_b, max_row_b = range_boundaries(range_str)
                    if (min_col <= c <= max_col_b) and (min_row <= source_row_idx <= max_row_b):
                        dv.ranges.add(target_range)
                        break

    @staticmethod
    def apply_static_content(config, source, destination, start_row, count):
        """Handles static values and formula translation for columns defined in config."""
        static_cols = 1
        while hasattr(config, f"column_static_{static_cols}_id"):
            col_id = getattr(config, f"column_static_{static_cols}_id")
            if hasattr(config, f"column_static_{static_cols}_value"):
                val = str(getattr(config, f"column_static_{static_cols}_value"))
                for row in range(start_row, start_row + count):
                    destination.cell(row=row, column=col_id).value = val.replace("{row}", str(row))
            else:
                formula = source.cell(row=start_row, column=col_id).value
                if formula:
                    origin = f"{get_column_letter(col_id)}{start_row}"
                    for row in range(start_row, start_row + count):
                        target = f"{get_column_letter(col_id)}{row}"
                        destination.cell(row=row, column=col_id).value = Translator(formula, origin=origin).translate_formula(target)
            static_cols += 1

    @staticmethod
    def populate_template_sheet(new_sheet, data_row, config, bom_name, component_id, vuln_id, vuln_desc, pub_date, vector_str, base_score, is_kev):
        """Writes standardized vulnerability data into the current row."""
        if hasattr(config, "column_id_bom"):
            new_sheet.cell(row=data_row, column=config.column_id_bom).value = bom_name
        if hasattr(config, "column_id_cpe"):
            new_sheet.cell(row=data_row, column=config.column_id_cpe).value = component_id
        if hasattr(config, "column_id_cve"):
            new_sheet.cell(row=data_row, column=config.column_id_cve).value = vuln_id
        
        if hasattr(config, "column_id_description"):
            desc = vuln_desc
            if hasattr(config, "template_max_description_char") and len(vuln_desc) > config.template_max_description_char:
                desc = f"{vuln_desc[:config.template_max_description_char]} (truncated)"
            new_sheet.cell(row=data_row, column=config.column_id_description).value = desc

        if hasattr(config, "column_id_publish_date"):
            new_sheet.cell(row=data_row, column=config.column_id_publish_date).value = pub_date[:10]
        if hasattr(config, "column_id_cvss"):
            new_sheet.cell(row=data_row, column=config.column_id_cvss).value = vector_str
        if hasattr(config, "column_id_base_score"):
            new_sheet.cell(row=data_row, column=config.column_id_base_score).value = base_score
        if hasattr(config, "column_id_is_kev"):
            new_sheet.cell(row=data_row, column=config.column_id_is_kev).value = is_kev

        # Optional CVSS Component Splitting
        if hasattr(config, "split_cvss_score") and config.split_cvss_score:
            try:
                tokens = CVSSHelper.tokenize_cvss3_human(vector_str)
                metric_map = {
                    "column_split_cvss_av": "AV", "column_split_cvss_ac": "AC",
                    "column_split_cvss_pr": "PR", "column_split_cvss_ui": "UI",
                    "column_split_cvss_s":  "S",  "column_split_cvss_c":  "C",
                    "column_split_cvss_i":  "I",  "column_split_cvss_a":  "A"
                }
                for cfg_attr, token_key in metric_map.items():
                    if hasattr(config, cfg_attr):
                        col = getattr(config, cfg_attr)
                        new_sheet.cell(row=data_row, column=col).value = tokens.get(token_key)
            except TypeError:
                pass

    @staticmethod
    def populate_epss_data(new_sheet, config, epss_manager):
        """Writes bulk EPSS scores to their previously registered rows."""
        if hasattr(config, "column_id_epss"):
            epss_col = config.column_id_epss
            for cve, epss_score, idx in epss_manager:
                new_sheet.cell(row=idx, column=epss_col).value = epss_score