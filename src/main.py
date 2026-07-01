#!/usr/bin/env python3
"""
Vulnerability Scanner Orchestrator
"""

import argparse
import os
import csv
import json
from pathlib import Path
from datetime import datetime

# Excel and Data Modeling
from openpyxl import load_workbook
from cyclonedx.model.bom import Bom

# Custom Modules
import config_handler
import core.nvd as nvd
import core.cisa as cisa
import core.osv as osv
import core.epss as epss
import core.dependency_track as dt
import core.aggregator as va
from helpers.excel_utilities import ExcelHelper

def get_component_list(file: str, file_type: str, csv_column_id: int):
    """
    Extracts software component identifiers (CPEs or PURLs) from a source file.

    Supported formats include flat CSV files (extracting from a specified column index) 
    and CycloneDX JSON Bill of Materials (SBOM) files.

    Args:
        file (str): Filepath to the inventory or SBOM document.
        file_type (str): The format of the file, either 'csv' or 'json' (CycloneDX).
        csv_column_id (int): The zero-indexed column position to parse if file_type is 'csv'.

    Returns:
        list[str]: A list of cleaned component identification strings.
    """
    cpe_list = []
    if file_type.lower() == "csv":
        with open(file, mode='r', encoding='utf-8') as f:
            for row in csv.reader(f):
                if len(row) > csv_column_id:
                    cpe_list.append(row[csv_column_id].strip())
    else:
        with open(file, 'r') as f:
            bom = Bom.from_json(data=json.load(f))
        for component in bom.components:
            if component.cpe:
                cpe_list.append(str(component.cpe))
            elif component.purl:
                cpe_list.append(str(component.purl))
    return cpe_list

def valid_date(s):
    """
    Validates and converts a command-line date string argument into a datetime object.

    Args:
        s (str): The input date string, expected in 'YYYY-MM-DD' format.

    Returns:
        datetime: A native datetime object representation of the input date.

    Raises:
        argparse.ArgumentTypeError: If the string does not match the expected format.
    """
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        raise argparse.ArgumentTypeError(f"Not a valid date: '{s}'. Expected format: YYYY-MM-DD.")

def _parse_published_date(v_data, verbose=False):
    """
    Safely extracts and parses the vulnerability publication date from a record.

    Converts ISO-8601 strings (handling 'Z' UTC suffixes) into timezone-naive 
    datetime objects for uniform comparison against date filters.

    Args:
        v_data (dict): A single vulnerability data dictionary containing a 'published' key.
        verbose (bool, optional): If True, logs date parsing failures to stdout. Defaults to False.

    Returns:
        datetime: The parsed publication date, or datetime.min if parsing fails.
    """
    try:
        return datetime.fromisoformat(v_data['published'].replace('Z', '+00:00')).replace(tzinfo=None)
    except (ValueError, KeyError):
        if verbose:
            print(f"    [!] Could not parse date for {v_data.get('cve_id')}, skipping date filter.")
        return datetime.min

def _apply_excel_formatting(sheet, template_sheet, config_template, row_count, mode_label):
    """
    Applies unified structural layouts, static components, and formatting rules to an Excel sheet.

    Wraps helper utilities to construct headers, copy styles from the base template, 
    map dynamic cell widths, and bind validation rules to user columns.

    Args:
        sheet (Worksheet): The target openpyxl sheet to style and format.
        template_sheet (Worksheet): The reference/source template worksheet.
        config_template (Namespace): Configuration section containing template row/column metrics.
        row_count (int): Total number of vulnerability rows populated in the sheet.
        mode_label (str): Text context for execution logs (e.g., "Individual SBOM", "Combined SBOM").
    """
    print(f"[*] Applying Excel Formatting and Content ({mode_label})...")
    start_row = config_template.template_start_row
    ExcelHelper.apply_static_content(config_template, template_sheet, sheet, start_row, row_count)
    ExcelHelper.apply_formatting_to_range(sheet, start_row, start_row + 1, row_count)
    ExcelHelper.apply_data_validation_rules(sheet, start_row, row_count)

def _main_sbom_file_processing(args, config, system_config, aggregator, epss_manager, kev_obj, wb, template_sheet, system_root_path):
    """
    Processes systems that reference local standalone SBOM files (CSVs or CycloneDX JSONs).

    Iterates through configured component manifests, queries aggregated threat engines 
    for records matching component markers, appends secondary risk models (EPSS and CISA KEV), 
    and constructs final localized workbooks/worksheets.

    Args:
        args (Namespace): Parsed command line parameters (filters, verbs, targets).
        config (NVDConfigFile): The application global settings configuration object.
        system_config (SystemConfigFile): Target system configuration specification object.
        aggregator (VulnerabilityAggregator): Unified API vulnerability aggregator object.
        epss_manager (EPSS): Initialized EPSS metrics query router.
        kev_obj (KEV): In-memory CISA Known Exploited Vulnerabilities catalog tracker.
        wb (Workbook): Active openpyxl workbook payload object.
        template_sheet (Worksheet): Active master template template worksheet.
        system_root_path (Path): Base operating directory path of the targeted system.
    """
    combine_sboms = getattr(system_config, "combine_all_boms", False)
    include_zero = getattr(config.GLOBAL, "include_zero_vuln_components", False)
    include_deferred = not getattr(config.GLOBAL, "ignore_defferred", False)
    template_name = getattr(config.TEMPLATE, "template_sheet_name", wb.sheetnames[0])

    new_sheet = template_sheet if combine_sboms else None
    data_row = config.TEMPLATE.template_start_row
    row_count = 0
    
    for bom in system_config.boms:
        full_bom = Path(system_root_path) / bom
        full_bom_name = full_bom.stem
        clean_name = full_bom_name[:31]

        if args.verbose:
            print(f"  [*] Analyzing SBOM: {full_bom.name}")

        if not combine_sboms:
            new_sheet = wb.copy_worksheet(template_sheet)
            if template_sheet.auto_filter.ref:
                new_sheet.auto_filter.ref = template_sheet.auto_filter.ref
            ExcelHelper.copy_data_validations(template_sheet, new_sheet)
            new_sheet.title = f"RA_{clean_name}"
            data_row = config.TEMPLATE.template_start_row
            row_count = 0

        csv_col = getattr(system_config, "bom_cpe_column", 0)
        component_list = get_component_list(str(full_bom), system_config.bom_format, csv_col)

        if args.verbose:
            print(f"  [+] Extracted {len(component_list)} components.")

        for component_id in component_list:
            clean_identifier = component_id.replace('\\', '')
            vulns_list = aggregator.get_vulnerabilities(clean_identifier)

            if not vulns_list:
                if include_zero:
                    ExcelHelper.populate_template_sheet(
                        new_sheet, data_row, config.TEMPLATE, full_bom_name, component_id, 
                        'None', 'No Vulnerabilities Found', 'N/A', 'N/A', 0, False
                    )
                    data_row += 1
                    row_count += 1
                continue

            if args.verbose:
                print(f"  [*] Queried {len(vulns_list)} for {clean_identifier}")

            for v_data in vulns_list:
                pub_dt = _parse_published_date(v_data, args.verbose)
                
                if (args.start and pub_dt < args.start) or (args.end and pub_dt > args.end):
                    continue
                if v_data.get('status') == "Deferred" and not include_deferred:
                    continue

                is_kev = kev_obj.query_cpe(v_data['cve_id']) if v_data['cve_id'].startswith("CVE-") else False
                if v_data['cve_id'].startswith("CVE-"):
                    epss_manager.register_cve(cve_id=v_data['cve_id'], indexer_id=data_row)

                ExcelHelper.populate_template_sheet(
                    new_sheet, data_row, config.TEMPLATE, full_bom_name, component_id,
                    v_data['cve_id'], v_data['description'], v_data['published'],
                    v_data['vector'], v_data['base_score'], is_kev
                )
                data_row += 1
                row_count += 1

        if row_count > 0:
            if args.verbose:
                print(f"  [*] Fetching bulk EPSS scores for {full_bom_name}...")
            epss_manager.query() 
            ExcelHelper.populate_epss_data(new_sheet, config.TEMPLATE, epss_manager)
            epss_manager.clear_registry()

        if not combine_sboms:
            _apply_excel_formatting(new_sheet, template_sheet, config.TEMPLATE, row_count, "Individual SBOM")
    
    if combine_sboms:
        _apply_excel_formatting(new_sheet, template_sheet, config.TEMPLATE, row_count, "Combined SBOM")
    else:
        if len(wb.sheetnames) > 1:
            wb.remove(wb[template_name])
        else:
            wb[template_name].title = "No_Vulnerabilities_Found"

def _main_sbom_uuid_processing(args, config, system_config, aggregator, epss_manager, kev_obj, template_sheet):
    """
    Processes systems integrated via remote server management instances (e.g., Dependency-Track).

    Leverages project/system UUID elements to collect systemic metrics natively computed 
    upstream instead of dissecting local files item by item. 

    Args:
        args (Namespace): Parsed command line parameters (filters, verbs, targets).
        config (NVDConfigFile): The application global settings configuration object.
        system_config (SystemConfigFile): Target system configuration specification object.
        aggregator (VulnerabilityAggregator): Unified API vulnerability aggregator object.
        epss_manager (EPSS): Initialized EPSS metrics query router.
        kev_obj (KEV): In-memory CISA Known Exploited Vulnerabilities catalog tracker.
        template_sheet (Worksheet): Active target destination sheet for vulnerability population.
    """
    system_uuid = getattr(system_config, "system_uuid", "")
    vulns_list = aggregator.get_vulnerabilities(system_uuid)

    if args.verbose:
        print(f"  [*] Queried {len(vulns_list)} for {system_uuid}")

    data_row = config.TEMPLATE.template_start_row
    row_count = 0

    for v_data in vulns_list:
        pub_dt = _parse_published_date(v_data, args.verbose)
        
        if (args.start and pub_dt < args.start) or (args.end and pub_dt > args.end):
            continue

        is_kev = kev_obj.query_cpe(v_data['cve_id']) if v_data['cve_id'].startswith("CVE-") else False
        if v_data['cve_id'].startswith("CVE-"):
            epss_manager.register_cve(cve_id=v_data['cve_id'], indexer_id=data_row)

        ExcelHelper.populate_template_sheet(
            template_sheet, data_row, config.TEMPLATE, system_uuid, "CPE",
            v_data['cve_id'], v_data['description'], v_data['published'],
            v_data['vector'], v_data['base_score'], is_kev
        )
        data_row += 1
        row_count += 1

    if row_count > 0:
        if args.verbose:
            print(f"  [*] Fetching bulk EPSS scores for {system_uuid}...")
        epss_manager.query() 
        ExcelHelper.populate_epss_data(template_sheet, config.TEMPLATE, epss_manager)
        epss_manager.clear_registry()

    _apply_excel_formatting(template_sheet, template_sheet, config.TEMPLATE, row_count, "Combined SBOM")

def main():
    """
    Application entry point. Coordinates command line setup, configuration bootstrapping, 
    risk database hydration, processing strategy dispatching, and final file outputs.
    """
    parser = argparse.ArgumentParser(
        description="Scan NVD/OSV for vulnerabilities associated with CPE/PURL inputs.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument("--config", type=str, default="config.ini", help="Path to main .ini config")
    parser.add_argument("--load", type=str, help="Load a pre-existing JSON dump for offline testing")
    parser.add_argument("--outdir", type=str, default="", help="Output directory for generated Excel files")
    parser.add_argument("--system_override", type=str, help="Specific System.ini to process")
    parser.add_argument("--start", type=valid_date, help="Filter vulnerabilities published AFTER this date")
    parser.add_argument("--end", type=valid_date, help="Filter vulnerabilities published BEFORE this date")
    parser.add_argument("--verbose", action="store_true", help="Enable detailed debug logging to console")
    args = parser.parse_args()

    if args.verbose:
        print(f"[*] Starting Orchestrator at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    config_file = Path(args.config).resolve()
    config_root = config_file.parent
    config = config_handler.NVDConfigFile(config_file)

    val_cert = getattr(config.GLOBAL, "validate_remote_certificate", True)
    score_ver = getattr(config.GLOBAL, "score_system_ver", "3.1")

    if args.verbose:
        setattr(config.GLOBAL, "verbose_logging", True)
        print("[*] Initializing API clients and loading KEV database...")    
        
    kev_obj = cisa.KEV(val_cert)
    kev_obj.load_kevs()
    epss_manager = epss.EPSS(verify_certificate=val_cert)
    aggregator = va.VulnerabilityAggregator(getattr(config.GLOBAL, "source_locations", "nvd,osv"), config, val_cert, score_ver)

    global_template_file = config_root / config.TEMPLATE.template
    if not global_template_file.exists():
        raise FileNotFoundError(f"Template not found: {global_template_file}")

    system_configs = [args.system_override] if args.system_override else config.GLOBAL.input_configs

    for system in system_configs:
        if args.verbose:
            print(f"\n>>> Processing System Config: {system}")
            
        system_config = config_handler.SystemConfigFile(system)
        system_root_path = Path(system).resolve().parent
        
        local_template_file = Path(system_root_path) / system_config.template if hasattr(system_config, "template") else global_template_file

        wb = load_workbook(filename=local_template_file, keep_vba=True)
        template_name = getattr(config.TEMPLATE, "template_sheet_name", wb.sheetnames[0])
        template_sheet = wb[template_name]

        override_sources = getattr(system_config, "source_locations", "")
        aggregator_override = va.VulnerabilityAggregator(override_sources, config, val_cert, score_ver) if override_sources else aggregator

        if aggregator_override.process_individual_cpes():
            _main_sbom_file_processing(args, config, system_config, aggregator_override, epss_manager, kev_obj, wb, template_sheet, system_root_path)
        else:
            print(f"[+] Using Dependency Track Logic so we have to use project IDs instead of CPEs")
            _main_sbom_uuid_processing(args, config, system_config, aggregator_override, epss_manager, kev_obj, template_sheet)
        
        out_file = Path(args.outdir) / f"{system_config.name[:31]}{Path(local_template_file).suffix}"
        wb.save(out_file)
        
        if args.verbose:
            print(f"[SUCCESS] Report generated: {out_file}")

if __name__ == "__main__":
    main()