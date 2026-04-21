import argparse
import openpyxl

def update_excel_data(source_path, dest_path, match_col, columns, match_col_idx, debug=False):
    """
    Synchronizes specific columns between Excel files by matching row keys.
    
    Args:
        source_path (str):
        dest_path (str):
        match_col (str):
        columns (list):
        match_col_idx (str):
    """
    wb_src = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    wb_dst = openpyxl.load_workbook(dest_path)
    
    total_updated = 0

    for sheet_name in wb_src.sheetnames:
        if sheet_name not in wb_dst.sheetnames:
            if debug: print(f"Skipping {sheet_name}: Not found in destination.")
            continue
        
        ws_src = wb_src[sheet_name]
        ws_dst = wb_dst[sheet_name]

        # Map destination: {str(key).strip(): row_index}
        dst_map = {}
        for row in range(1, ws_dst.max_row + 1):
            val = ws_dst[f"{match_col_idx}{row}"].value
            if val is not None:
                dst_map[str(val).strip()] = row

        for row in range(1, ws_src.max_row + 1):
            raw_key = ws_src[f"{match_col_idx}{row}"].value
            key = str(raw_key).strip() if raw_key is not None else None
            
            if key in dst_map:
                dest_row = dst_map[key]
                for col in columns:
                    ws_dst[f"{col}{dest_row}"] = ws_src[f"{col}{row}"].value
                total_updated += 1
    
    wb_dst.save(dest_path)
    return total_updated


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sync Excel columns row-by-row.")
    parser.add_argument("--source", required=True, help="Original risk assessment containing previous answers and content")
    parser.add_argument("--destination", required=True, help="New excel file containing old and new vulnerabilities")
    parser.add_argument("--columns", nargs='+', required=True, help="List of columns to copy from source to destination if a match is found")
    parser.add_argument("--match", dest="match_col_idx", required=True, help="Column letter to use as the matching key")
    parser.add_argument("--debug", required=False, action="store_true", help="Column letter to use as the matching key")

    args = parser.parse_args()
    updated_count = update_excel_data(args.source, args.destination, None, args.columns, args.match_col_idx)
    print(f"Successfully synced {updated_count} rows to {args.destination}")

